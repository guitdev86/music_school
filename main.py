import openpyxl as xl
#from openpyxl.chart import BarChart, Reference

class Accounting:
    def __init__(self, filename):
        self.filename = filename
        self.wb = xl.load_workbook(filename)
        self.sheet = self.wb['Sheet1']

    def calculate_earnngs(self, filename):
        overall_earnings = 0
        for row in range(3, self.sheet.max_row + 1):
            cell = self.sheet.cell(row, 3)
            overall_earnings += cell.value
        return overall_earnings

    def add_payment(self, name, payment, date):
        self.sheet.cell(self.sheet.max_row + 1, 1).value = name
        self.sheet.cell(self.sheet.max_row, 3).value = date
        self.sheet.cell(self.sheet.max_row, 2).value = payment
        self.wb.save(self.filename)
        return

    def find_payment(self, name):
        total = 0
        for row in range(1, self.sheet.max_row + 1):
            cell = self.sheet.cell(row, 1).value
            if cell == name:
                amount = self.sheet.cell(row, 3).value
                date = self.sheet.cell(row, 2).value
                total += amount
                print(f'The amount payed by {cell} on {date} is: {amount}')
        return print(f'The total amount payed by {name} is: {total}')

    def get_attendance(self):
        pass

    def get_balance(self):
        pass

obj = Accounting('source.xlsx')
#obj.add_payment('Anna', '10.09.2019', 15)
obj.find_payment('Maksim')